from flask import Flask, request, send_file
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.drawing.image import Image as OpenpyxlImage
import io
import os
from datetime import datetime
from copy import copy

app = Flask(__name__)

# Path to your SF2 template (place in same folder as this file)
TEMPLATE_PATH = 'SF2_template.xlsx'

def get_month_index(month_name):
    """Convert month name to index"""
    months = ['January', 'February', 'March', 'April', 'May', 'June',
              'July', 'August', 'September', 'October', 'November', 'December']
    try:
        return months.index(month_name) + 1
    except ValueError:
        return 0

def get_weekdays_in_month(year, month_index):
    """Get list of weekday dates (Mon-Fri) in the given month"""
    from datetime import datetime, timedelta
    
    # Get first day of month
    first_day = datetime(year, month_index, 1)
    
    # Get last day of month
    if month_index == 12:
        last_day = datetime(year + 1, 1, 1) - timedelta(days=1)
    else:
        last_day = datetime(year, month_index + 1, 1) - timedelta(days=1)
    
    weekdays = []
    current_date = first_day
    
    while current_date <= last_day:
        # weekday() returns 0-6, where 0=Monday, 4=Friday, 5=Saturday, 6=Sunday
        if current_date.weekday() < 5:  # Monday to Friday
            weekdays.append(current_date)
        current_date += timedelta(days=1)
    
    return weekdays

@app.route('/api/generate-sf2', methods=['POST'])
def generate_sf2():
    """Generate SF2 report from attendance data (weekdays only)"""
    try:
        # Get data from request
        data = request.json
        month = data.get('month')
        year = data.get('year')
        students = data.get('students', [])
        
        print(f"\n{'='*60}")
        print(f"Generating SF2 for {month} {year}")
        print(f"Processing {len(students)} students")
        print(f"{'='*60}\n")
        
        # Validate month
        month_index = get_month_index(month)
        if month_index == 0:
            return {'error': f'Invalid month: {month}'}, 400
        
        # Check if template exists
        if not os.path.exists(TEMPLATE_PATH):
            return {'error': f'Template not found: {TEMPLATE_PATH}'}, 500
        
        # Load template
        wb = load_workbook(TEMPLATE_PATH)
        ws = wb.active
        
        print(f"Template loaded: {ws.title}")
        
        # PRESERVE IMAGES - Deep copy approach
        saved_images = []
        if hasattr(ws, '_images'):
            print(f"Found {len(ws._images)} images in template")
            for img in ws._images:
                try:
                    # Create a new image object from the existing one
                    new_img = OpenpyxlImage(img.ref)
                    new_img.anchor = copy(img.anchor)
                    new_img.width = img.width
                    new_img.height = img.height
                    saved_images.append(new_img)
                    print(f"  Image anchored at: {img.anchor}")
                except Exception as e:
                    print(f"  Warning: Could not copy image: {e}")
        
        # Store merged cells to restore later
        merged_cells = list(ws.merged_cells.ranges)
        print(f"Found {len(merged_cells)} merged cell ranges\n")
        
        # UNMERGE ALL CELLS temporarily
        for merged_range in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(merged_range))
        
        print(f"✅ All cells unmerged\n")
        
        # Write month/year to X6
        ws['X6'] = f"{month} {year}"
        print(f"✅ Written month/year to X6: {month} {year}\n")
        
        # Get weekdays in month
        weekdays = get_weekdays_in_month(year, month_index)
        num_weekdays = len(weekdays)
        
        print(f"📅 Weekdays in month: {num_weekdays}")
        print(f"First weekday: {weekdays[0].strftime('%Y-%m-%d (%A)')}")
        print(f"Last weekday: {weekdays[-1].strftime('%Y-%m-%d (%A)')}\n")
        
        # Day names mapping
        day_names = ['M', 'T', 'W', 'Th', 'F', 'Sa', 'Su']
        
        # Write day numbers and day names ONLY for weekdays
        print("Writing day headers:")
        for idx, weekday_date in enumerate(weekdays):
            day = weekday_date.day
            day_of_week = weekday_date.weekday()  # 0=Monday, 4=Friday
            day_name = day_names[day_of_week]
            
            # Column: D=4, E=5, F=6, etc. (openpyxl uses 1-based indexing)
            col_num = 4 + idx
            
            # Write day number to row 11
            ws.cell(row=11, column=col_num).value = day
            
            # Write day name to row 12
            ws.cell(row=12, column=col_num).value = day_name
            
            if idx < 5 or idx >= num_weekdays - 2:  # Show first 5 and last 2
                print(f"  Day {day:2d} ({day_name:2s}) -> Row 11-12, Column {col_num} (Excel col {chr(64+col_num)})")
        
        print(f"✅ Written {num_weekdays} weekday numbers to row 11")
        print(f"✅ Written {num_weekdays} weekday names to row 12\n")
        
        # Separate students by gender
        male_students = [s for s in students if s.get('gender', '').upper() == 'MALE']
        female_students = [s for s in students if s.get('gender', '').upper() == 'FEMALE']
        
        # Sort alphabetically
        male_students.sort(key=lambda s: s.get('name', ''))
        female_students.sort(key=lambda s: s.get('name', ''))
        
        print(f"Males: {len(male_students)}, Females: {len(female_students)}\n")
        
        # Define row constants
        male_start_row = 14
        female_start_row = 36
        
        # Create a mapping of day number to column (for weekdays only)
        day_to_col = {}
        for idx, weekday_date in enumerate(weekdays):
            day_to_col[weekday_date.day] = 4 + idx
        
        print(f"Day-to-Column mapping created: {len(day_to_col)} days\n")
        
        # Write MALE students (rows 14-34)
        print("Processing MALE students:")
        for idx, student in enumerate(male_students[:21]):  # Max 21 male students
            row = male_start_row + idx
            
            # Write row number in column A
            ws.cell(row=row, column=1).value = idx + 1
            
            # Write student name in column B
            ws.cell(row=row, column=2).value = student.get('name', '')
            
            print(f"  Male #{idx+1:2d}: {student.get('name', ''):<30s} at row {row}")
            
            # Initialize counters for this student
            absent_count = 0
            tardy_count = 0  # Late + Cutting Class combined
            
            # Write attendance
            attendance = student.get('attendance', [])
            for att in attendance:
                try:
                    att_date = datetime.strptime(att.get('date'), '%Y-%m-%d')
                    
                    # Check if date is in current month/year AND is a weekday
                    if att_date.month == month_index and att_date.year == year:
                        day = att_date.day
                        
                        # Only process if this day is in our weekdays list
                        if day in day_to_col:
                            col_num = day_to_col[day]
                            cell = ws.cell(row=row, column=col_num)
                            
                            status = att.get('status', '')
                            if status == 'Absent':
                                cell.value = 'x'
                                cell.font = Font(bold=True, color='000000')
                                absent_count += 1
                            elif status == 'Late':
                                cell.value = ''
                                cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                                tardy_count += 1
                            elif status == 'Cutting Class':
                                cell.value = ''
                                cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                                tardy_count += 1
                except Exception as e:
                    print(f"    Error processing attendance: {e}")
            
            # Write counts to columns AC (29) and AD (30)
            ws.cell(row=row, column=29).value = absent_count
            ws.cell(row=row, column=30).value = tardy_count
            print(f"             Absent: {absent_count}, Tardy: {tardy_count}")
        
        print()
        
        # Calculate daily present count for MALES (row 35)
        total_male_students = len(male_students)
        print(f"Calculating daily male attendance (Total enrolled: {total_male_students}):")
        
        for idx in range(num_weekdays):
            col_num = 4 + idx
            
            # Count absent and tardy for this day among males
            absent_tardy_count = 0
            for student_idx in range(total_male_students):
                student_row = male_start_row + student_idx
                cell = ws.cell(row=student_row, column=col_num)
                
                # Check if cell has 'x' (absent) or has fill (tardy)
                if cell.value == 'x' or (cell.fill and cell.fill.start_color and 
                                        cell.fill.start_color.rgb and 
                                        cell.fill.start_color.rgb != '00000000'):
                    absent_tardy_count += 1
            
            # Present = Total - (Absent + Tardy)
            present_count = total_male_students - absent_tardy_count
            ws.cell(row=35, column=col_num).value = present_count
            
            if idx < 5:  # Show first 5 days
                print(f"  Day idx {idx}, Col {col_num}: {present_count} present")
        
        print(f"✅ Daily male present counts written to row 35\n")
        
        # Write FEMALE students (rows 36-60)
        print("Processing FEMALE students:")
        for idx, student in enumerate(female_students[:25]):  # Max 25 female students
            row = female_start_row + idx
            
            # Write row number in column A
            ws.cell(row=row, column=1).value = idx + 1
            
            # Write student name in column B
            ws.cell(row=row, column=2).value = student.get('name', '')
            
            print(f"  Female #{idx+1:2d}: {student.get('name', ''):<30s} at row {row}")
            
            # Initialize counters for this student
            absent_count = 0
            tardy_count = 0
            
            # Write attendance
            attendance = student.get('attendance', [])
            for att in attendance:
                try:
                    att_date = datetime.strptime(att.get('date'), '%Y-%m-%d')
                    
                    # Check if date is in current month/year AND is a weekday
                    if att_date.month == month_index and att_date.year == year:
                        day = att_date.day
                        
                        # Only process if this day is in our weekdays list
                        if day in day_to_col:
                            col_num = day_to_col[day]
                            cell = ws.cell(row=row, column=col_num)
                            
                            status = att.get('status', '')
                            if status == 'Absent':
                                cell.value = 'x'
                                cell.font = Font(bold=True, color='000000')
                                absent_count += 1
                            elif status == 'Late':
                                cell.value = ''
                                cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                                tardy_count += 1
                            elif status == 'Cutting Class':
                                cell.value = ''
                                cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                                tardy_count += 1
                except Exception as e:
                    print(f"    Error processing attendance: {e}")
            
            # Write counts to columns AC (29) and AD (30)
            ws.cell(row=row, column=29).value = absent_count
            ws.cell(row=row, column=30).value = tardy_count
            print(f"             Absent: {absent_count}, Tardy: {tardy_count}")
        
        print()
        
        # Calculate daily present count for FEMALES (row 61)
        total_female_students = len(female_students)
        print(f"Calculating daily female attendance (Total enrolled: {total_female_students}):")
        
        for idx in range(num_weekdays):
            col_num = 4 + idx
            
            # Count absent and tardy for this day among females
            absent_tardy_count = 0
            for student_idx in range(total_female_students):
                student_row = female_start_row + student_idx
                cell = ws.cell(row=student_row, column=col_num)
                
                # Check if cell has 'x' (absent) or has fill (tardy)
                if cell.value == 'x' or (cell.fill and cell.fill.start_color and 
                                        cell.fill.start_color.rgb and 
                                        cell.fill.start_color.rgb != '00000000'):
                    absent_tardy_count += 1
            
            # Present = Total - (Absent + Tardy)
            present_count = total_female_students - absent_tardy_count
            ws.cell(row=61, column=col_num).value = present_count
        
        print(f"✅ Daily female present counts written to row 61\n")
        
        # Calculate daily TOTAL present count (male + female) (row 62)
        print("Calculating daily total attendance:")
        for idx in range(num_weekdays):
            col_num = 4 + idx
            
            # Get male and female present counts for this day
            male_present = ws.cell(row=35, column=col_num).value or 0
            female_present = ws.cell(row=61, column=col_num).value or 0
            
            # Total present
            total_present = male_present + female_present
            ws.cell(row=62, column=col_num).value = total_present
            
            if idx < 5:  # Show first 5 days
                print(f"  Day idx {idx}, Col {col_num}: M={male_present} + F={female_present} = {total_present}")
        
        print(f"✅ Daily total present counts written to row 62\n")
        
        # Re-merge ALL cells to restore template formatting
        print(f"Re-merging {len(merged_cells)} cell ranges...")
        merge_count = 0
        for merged_range in merged_cells:
            try:
                ws.merge_cells(str(merged_range))
                merge_count += 1
            except Exception as e:
                print(f"  Warning: Could not re-merge {merged_range}: {e}")
        
        print(f"✅ Successfully re-merged {merge_count}/{len(merged_cells)} cell ranges\n")
        
        # Restore all images
        if saved_images:
            print(f"Restoring {len(saved_images)} images...")
            # Clear existing images
            ws._images = []
            
            # Add back all saved images
            for idx, img in enumerate(saved_images):
                try:
                    ws.add_image(img)
                    print(f"  ✓ Image {idx+1} restored at {img.anchor}")
                except Exception as e:
                    print(f"  ✗ Failed to restore image {idx+1}: {e}")
            
            print(f"✅ Images restoration complete\n")
        else:
            print("⚠️ No images found in template\n")
        
        # Save to memory
        print("Saving workbook to memory...")
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        print(f"\n{'='*60}")
        print(f"✅ SF2 report generated successfully")
        print(f"{'='*60}\n")
        
        # Return file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"SF2_{month}_{year}.xlsx"
        )
    
    except Exception as e:
        print(f"\n{'='*60}")
        print(f"❌ ERROR: {str(e)}")
        print(f"{'='*60}\n")
        import traceback
        traceback.print_exc()
        return {'error': str(e)}, 500

@app.route('/health', methods=['GET'])
def health():
    """Health check endpoint"""
    return {'status': 'ok', 'timestamp': datetime.now().isoformat()}

if __name__ == '__main__':
    print("\n" + "="*60)
    print("Starting SF2 Backend Server...")
    print(f"Template file: {TEMPLATE_PATH}")
    print("="*60 + "\n")
    app.run(debug=True, host='0.0.0.0', port=5000)
